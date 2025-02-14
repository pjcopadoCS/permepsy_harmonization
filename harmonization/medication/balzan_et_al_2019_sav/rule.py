import os
import pandas as pd
from openpyxl import load_workbook



def harmonize(dataset,position,medication):


    base_folder = 'harmonization'
    folder = 'input_db'
    file = 'medication_cabecera_UV.xlsx'
    # Load the Excel workbook
    workbook = load_workbook(os.sep.join([base_folder,folder,file]))

    # Select the active worksheet
    worksheet = workbook.active
    
    # Iterate through cells and get their colors
    for idx,row in enumerate(worksheet.iter_rows()):
        
        if position[0]+1 <= idx <= position[1]: 

            for pos,cell in enumerate(row):
                if cell.value in medication:
                    med = cell.value
                    if medication[med] == 'oral':
                        dose = row[pos]+1
                        print(md,dose)
                '''color = cell.font.color.rgb
                med = str(cell.value).strip().lower()
                if color == FFFFC000 or color == FFFF9900:
                    if med not in medication:
                        medication[med] = oral
                elif color ==FF70AD47:
                    if med not in medication:
                        medication[cell.value] = LAI'''